import io
import os
import unittest
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient

from .models import DeliverySession, DeliveryStop, SharedRoute, UserProfile
from .parsers import parse_csv, parse_file, parse_txt, parse_xlsx, parse_xml
from .serializers import ActiveSessionSerializer, SessionListSerializer

# ============================================
# Helpers
# ============================================


def _make_biker(username="biker1"):
    """Create a biker user with profile and token. Returns (user, client)."""
    user = User.objects.create_user(username=username)
    UserProfile.objects.create(user=user, role="biker")
    token = Token.objects.create(user=user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")
    return user, client


def _make_planner(username="planner1"):
    """Create a planner user with profile and token. Returns (user, client)."""
    user = User.objects.create_user(username=username)
    UserProfile.objects.create(user=user, role="planner")
    token = Token.objects.create(user=user)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")
    return user, client


def _make_optimized_session(owner, num_stops=3):
    """Create a session with geocoded, optimized stops ready to start."""
    session = DeliverySession.objects.create(owner=owner, name="Test Route", status="not_started")
    stops = []
    for i in range(num_stops):
        stops.append(
            DeliveryStop(
                session=session,
                name=f"Stop {i + 1}",
                raw_address=f"Address {i + 1}",
                lat=47.5 + i * 0.01,
                lng=19.08 + i * 0.01,
                geocode_status="success",
                sequence_order=i + 1,
            )
        )
    DeliveryStop.objects.bulk_create(stops)
    return session


# ============================================
# Priority 1: Route Lifecycle Tests
# ============================================


class RouteLifecycleTest(TestCase):
    def setUp(self):
        self.user, self.client = _make_biker("route_biker")
        self.session = _make_optimized_session(self.user, num_stops=3)

    def test_start_route_happy_path(self):
        response = self.client.patch(f"/api/sessions/{self.session.id}/start/")
        self.assertEqual(response.status_code, 200)
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, "in_progress")
        self.assertIsNotNone(self.session.started_at)
        self.assertEqual(self.session.current_stop_index, 1)

    def test_start_route_already_started(self):
        self.session.status = "in_progress"
        self.session.save(update_fields=["status"])
        response = self.client.patch(f"/api/sessions/{self.session.id}/start/")
        self.assertEqual(response.status_code, 400)

    def test_start_route_not_optimized(self):
        self.session.stops.update(sequence_order=None)
        response = self.client.patch(f"/api/sessions/{self.session.id}/start/")
        self.assertEqual(response.status_code, 400)

    def test_start_route_session_not_found(self):
        response = self.client.patch("/api/sessions/00000000-0000-0000-0000-000000000000/start/")
        self.assertEqual(response.status_code, 404)

    def test_update_stop_delivered(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        stop = self.session.stops.get(sequence_order=1)
        response = self.client.patch(
            f"/api/sessions/{self.session.id}/stops/{stop.id}/status/",
            {"status": "delivered"},
        )
        self.assertEqual(response.status_code, 200)
        stop.refresh_from_db()
        self.assertEqual(stop.delivery_status, "delivered")

    def test_update_stop_not_received(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        stop = self.session.stops.get(sequence_order=1)
        response = self.client.patch(
            f"/api/sessions/{self.session.id}/stops/{stop.id}/status/",
            {"status": "not_received"},
        )
        self.assertEqual(response.status_code, 200)
        stop.refresh_from_db()
        self.assertEqual(stop.delivery_status, "not_received")

    def test_update_stop_skipped(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        stop = self.session.stops.get(sequence_order=1)
        response = self.client.patch(
            f"/api/sessions/{self.session.id}/stops/{stop.id}/status/",
            {"status": "skipped"},
        )
        self.assertEqual(response.status_code, 200)
        stop.refresh_from_db()
        self.assertEqual(stop.delivery_status, "skipped")

    def test_update_stop_invalid_status(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        stop = self.session.stops.get(sequence_order=1)
        response = self.client.patch(
            f"/api/sessions/{self.session.id}/stops/{stop.id}/status/",
            {"status": "lost_in_space"},
        )
        self.assertEqual(response.status_code, 400)

    def test_update_stop_auto_advance(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        stop1 = self.session.stops.get(sequence_order=1)
        self.client.patch(
            f"/api/sessions/{self.session.id}/stops/{stop1.id}/status/",
            {"status": "delivered"},
        )
        self.session.refresh_from_db()
        self.assertEqual(self.session.current_stop_index, 2)

    def test_update_stop_auto_finish(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        for stop in self.session.stops.order_by("sequence_order"):
            self.client.patch(
                f"/api/sessions/{self.session.id}/stops/{stop.id}/status/",
                {"status": "delivered"},
            )
        self.session.refresh_from_db()
        self.assertEqual(self.session.status, "finished")
        self.assertIsNotNone(self.session.finished_at)
        self.assertIsNone(self.session.current_stop_index)

    def test_update_stop_not_found(self):
        self.client.patch(f"/api/sessions/{self.session.id}/start/")
        response = self.client.patch(
            f"/api/sessions/{self.session.id}/stops/999999/status/",
            {"status": "delivered"},
        )
        self.assertEqual(response.status_code, 404)


# ============================================
# Priority 2: Auth & Role Enforcement Tests
# ============================================


class AuthEndpointTest(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_login_creates_user_and_token(self):
        response = self.client.post("/api/auth/login/", {"username": "newbiker", "role": "biker"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("token", data)
        self.assertEqual(data["user"]["username"], "newbiker")
        self.assertEqual(data["user"]["role"], "biker")
        self.assertTrue(User.objects.filter(username="newbiker").exists())

    def test_login_existing_user(self):
        self.client.post("/api/auth/login/", {"username": "existuser", "role": "biker"})
        response = self.client.post("/api/auth/login/", {"username": "existuser", "role": "biker"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(User.objects.filter(username="existuser").count(), 1)

    def test_login_role_switch(self):
        self.client.post("/api/auth/login/", {"username": "switcher", "role": "biker"})
        self.client.post("/api/auth/login/", {"username": "switcher", "role": "planner"})
        profile = UserProfile.objects.get(user__username="switcher")
        self.assertEqual(profile.role, "planner")

    def test_login_empty_username(self):
        response = self.client.post("/api/auth/login/", {"username": "", "role": "biker"})
        self.assertEqual(response.status_code, 400)

    def test_login_invalid_role(self):
        response = self.client.post("/api/auth/login/", {"username": "baduser", "role": "admin"})
        self.assertEqual(response.status_code, 400)

    def test_me_returns_user(self):
        user, client = _make_biker("meuser")
        response = client.get("/api/auth/me/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["username"], "meuser")
        self.assertEqual(response.json()["role"], "biker")

    def test_me_unauthenticated(self):
        response = self.client.get("/api/auth/me/")
        self.assertEqual(response.status_code, 401)

    def test_logout_deletes_token(self):
        user, client = _make_biker("logoutuser")
        response = client.post("/api/auth/logout/")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Token.objects.filter(user=user).exists())


class RoleEnforcementTest(TestCase):
    def setUp(self):
        self.biker_a, self.client_a = _make_biker("rbac_biker_a")
        self.biker_b, self.client_b = _make_biker("rbac_biker_b")
        self.planner, self.planner_client = _make_planner("rbac_planner")

    def test_biker_cannot_access_other_bikers_session(self):
        session = DeliverySession.objects.create(owner=self.biker_a, name="A's route")
        response = self.client_b.get(f"/api/sessions/{session.id}/")
        self.assertEqual(response.status_code, 404)

    def test_planner_can_access_any_session(self):
        session = DeliverySession.objects.create(owner=self.biker_a, name="A's route")
        response = self.planner_client.get(f"/api/sessions/{session.id}/")
        self.assertEqual(response.status_code, 200)

    def test_unauthenticated_upload_rejected(self):
        client = APIClient()
        f = SimpleUploadedFile("stops.csv", b"name,address\nShop,Main St\n", content_type="text/csv")
        response = client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 401)

    def test_unauthenticated_sessions_rejected(self):
        client = APIClient()
        response = client.get("/api/sessions/")
        self.assertEqual(response.status_code, 401)


# ============================================
# Priority 3: Planner Management Tests
# ============================================


class PlannerManagementTest(TestCase):
    def setUp(self):
        self.planner, self.planner_client = _make_planner("mgmt_planner")
        self.biker, self.biker_client = _make_biker("mgmt_biker")
        self.session = DeliverySession.objects.create(owner=self.biker, name="Biker Route")

    def test_planner_delete_session(self):
        response = self.planner_client.delete(f"/api/sessions/{self.session.id}/delete/")
        self.assertEqual(response.status_code, 204)
        self.assertFalse(DeliverySession.objects.filter(id=self.session.id).exists())

    def test_biker_cannot_delete_session(self):
        response = self.biker_client.delete(f"/api/sessions/{self.session.id}/delete/")
        self.assertEqual(response.status_code, 403)

    def test_delete_session_not_found(self):
        response = self.planner_client.delete("/api/sessions/00000000-0000-0000-0000-000000000000/delete/")
        self.assertEqual(response.status_code, 404)

    def test_planner_assign_session(self):
        biker2, _ = _make_biker("assign_target")
        response = self.planner_client.patch(
            f"/api/sessions/{self.session.id}/assign/",
            {"owner_id": biker2.id},
        )
        self.assertEqual(response.status_code, 200)
        self.session.refresh_from_db()
        self.assertEqual(self.session.owner, biker2)

    def test_unassign_session(self):
        """Sending no owner_id unassigns the session."""
        response = self.planner_client.patch(f"/api/sessions/{self.session.id}/assign/", {})
        self.assertEqual(response.status_code, 200)
        self.session.refresh_from_db()
        self.assertIsNone(self.session.owner)

    def test_assign_user_not_found(self):
        response = self.planner_client.patch(
            f"/api/sessions/{self.session.id}/assign/",
            {"owner_id": 99999},
        )
        self.assertEqual(response.status_code, 404)

    def test_biker_cannot_assign(self):
        response = self.biker_client.patch(
            f"/api/sessions/{self.session.id}/assign/",
            {"owner_id": self.biker.id},
        )
        self.assertEqual(response.status_code, 403)

    def test_planner_rename_session(self):
        response = self.planner_client.patch(
            f"/api/sessions/{self.session.id}/rename/",
            {"name": "New Name"},
        )
        self.assertEqual(response.status_code, 200)
        self.session.refresh_from_db()
        self.assertEqual(self.session.name, "New Name")

    def test_rename_empty_name(self):
        response = self.planner_client.patch(
            f"/api/sessions/{self.session.id}/rename/",
            {"name": ""},
        )
        self.assertEqual(response.status_code, 400)

    def test_biker_cannot_rename(self):
        response = self.biker_client.patch(
            f"/api/sessions/{self.session.id}/rename/",
            {"name": "Hacked"},
        )
        self.assertEqual(response.status_code, 403)


class SessionListingTest(TestCase):
    def setUp(self):
        self.biker_a, self.client_a = _make_biker("list_biker_a")
        self.biker_b, self.client_b = _make_biker("list_biker_b")
        self.planner, self.planner_client = _make_planner("list_planner")
        DeliverySession.objects.create(owner=self.biker_a, name="A1")
        DeliverySession.objects.create(owner=self.biker_a, name="A2")
        DeliverySession.objects.create(owner=self.biker_b, name="B1")

    def test_biker_sees_only_own_sessions(self):
        response = self.client_a.get("/api/sessions/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 2)

    def test_planner_sees_all_sessions(self):
        response = self.planner_client.get("/api/sessions/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 3)

    def test_planner_filter_by_owner_id(self):
        response = self.planner_client.get(f"/api/sessions/?owner_id={self.biker_a.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 2)


# ============================================
# Priority 4: Sharing Tests
# ============================================


class SharingTest(TestCase):
    def setUp(self):
        self.user, self.client = _make_biker("share_biker")
        self.session = DeliverySession.objects.create(owner=self.user, name="Share Route")

    def test_create_share_link(self):
        response = self.client.post(f"/api/sessions/{self.session.id}/share/")
        self.assertEqual(response.status_code, 201)
        self.assertIn("share_id", response.json())

    def test_get_shared_route(self):
        share = SharedRoute.objects.create(session=self.session)
        anon_client = APIClient()
        response = anon_client.get(f"/api/shared/{share.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["session"]["name"], "Share Route")

    def test_shared_route_not_found(self):
        anon_client = APIClient()
        response = anon_client.get("/api/shared/00000000-0000-0000-0000-000000000000/")
        self.assertEqual(response.status_code, 404)

    def test_share_session_not_found(self):
        response = self.client.post("/api/sessions/00000000-0000-0000-0000-000000000000/share/")
        self.assertEqual(response.status_code, 404)


# ============================================
# Priority 5: XLSX Parsing Tests
# ============================================


class ParserXLSXTest(TestCase):
    def _make_xlsx(self, headers, data_rows):
        """Create an in-memory XLSX file."""
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_parse_xlsx_basic(self):
        f = self._make_xlsx(["name", "address"], [["Shop A", "Main St"], ["Shop B", "Vaci ut"]])
        rows = parse_xlsx(f)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Shop A")
        self.assertEqual(rows[0]["address"], "Main St")

    def test_parse_xlsx_with_delivery_details(self):
        f = self._make_xlsx(
            ["name", "address", "product_code", "recipient_name", "recipient_phone"],
            [["Shop A", "Main St", "PKG-001", "John Doe", "+36301234567"]],
        )
        rows = parse_xlsx(f)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product_code"], "PKG-001")
        self.assertEqual(rows[0]["recipient_name"], "John Doe")
        self.assertEqual(rows[0]["recipient_phone"], "+36301234567")


class ParserUPSSignatureTest(TestCase):
    """Parsing the real-world Hungarian UPS courier route sheet signature.

    Mirrors `example_files/UPS térkép teszt.xlsx` (gitignored company data):
    Hungarian headers, a per-courier summary count row, a numeric postal code,
    an address column built from an uncalculated Excel formula, and notes that
    stand in for the stop name.
    """

    # Headers exactly as in the source file (note the trailing space on "u ").
    HEADERS = ["város", "irszám", "u ", "hsz", "megj", "Cím", "kör", "bálint", "marci"]

    def _make_ups_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Munkalap1"
        ws.append(self.HEADERS)
        # Summary count row: only the courier columns carry values.
        ws.append([None] * 7 + [4, 3])
        # Data rows. Cím is an uncalculated formula -> parser must recompose it
        # from the street (col C) + house number (col D) columns.
        ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "Budapest", 1011, "Pala utca", "6"
        ws["E3"], ws["F3"], ws["G3"] = None, '=C3&" "&D3', "bálint"
        ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Budapest", 1011, "Markovits Iván utca", "4"
        ws["E4"], ws["F4"], ws["G4"] = "coyote cafe laverde", '=C4&" "&D4', "bálint"
        ws["A5"], ws["B5"], ws["C5"], ws["D5"] = "Budapest", 1117, "Gábor Dénes utca", "4"
        ws["E5"], ws["F5"], ws["G5"] = "icenter", '=C5&" "&D5', "fel"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_summary_row_dropped_real_stops_kept(self):
        rows = parse_xlsx(self._make_ups_xlsx())
        # The summary count row carries no address, so it is filtered out.
        self.assertEqual(len(rows), 3)

    def test_address_recomposed_from_components_with_locality(self):
        rows = parse_xlsx(self._make_ups_xlsx())
        # Uncalculated "=C3&..." formula must not leak; address is rebuilt from
        # street + house number, with the numeric postal code and city appended
        # (no trailing ".0" on the postal code).
        addr = rows[0]["address"]
        self.assertEqual(addr, "Pala utca 6, 1011 Budapest")
        self.assertFalse(any(r["address"].startswith("=") for r in rows))
        self.assertFalse(any(".0" in r["address"] for r in rows))

    def test_note_becomes_name_else_address(self):
        rows = parse_xlsx(self._make_ups_xlsx())
        by_addr = {r["address"]: r for r in rows}
        # Row with a megj note uses it as the stop name...
        self.assertEqual(
            by_addr["Markovits Iván utca 4, 1011 Budapest"]["name"],
            "coyote cafe laverde",
        )
        # ...and a row with no note falls back to the address as the name.
        self.assertEqual(
            by_addr["Pala utca 6, 1011 Budapest"]["name"],
            "Pala utca 6, 1011 Budapest",
        )

    def test_committed_fixture_matches_signature(self):
        """The committed e2e fixture parses cleanly to its 14 courier stops."""
        fixture = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "frontend",
            "e2e",
            "test-data",
            "ups_terkep_teszt.xlsx",
        )
        with open(fixture, "rb") as f:
            rows = parse_xlsx(f)
        self.assertEqual(len(rows), 14)
        self.assertTrue(all(r["name"] and r["address"] for r in rows))
        self.assertTrue(all("Budapest" in r["address"] for r in rows))

    @unittest.skipUnless(
        os.path.exists(
            os.path.join(
                os.path.dirname(__file__),
                "..",
                "..",
                "example_files",
                "UPS térkép teszt.xlsx",
            )
        ),
        "real company file is gitignored / not present",
    )
    def test_real_company_file_when_present(self):
        """When the real (gitignored) file is present locally, it parses cleanly."""
        path = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "example_files",
            "UPS térkép teszt.xlsx",
        )
        with open(path, "rb") as f:
            rows = parse_xlsx(f)
        self.assertGreater(len(rows), 100)
        self.assertTrue(all(r["name"] and r["address"] for r in rows))
        self.assertFalse(any(r["address"].startswith("=") for r in rows))


# ============================================
# Priority 6: Geocoder & Optimizer Mocked Tests
# ============================================


class GeocoderTest(TestCase):
    @patch("planner.geocoder.time")
    @patch("planner.geocoder.requests.get")
    def test_geocode_success(self, mock_get, mock_time):
        mock_time.monotonic.return_value = 100.0
        mock_response = MagicMock()
        mock_response.json.return_value = [{"lat": "47.5", "lon": "19.08"}]
        mock_response.raise_for_status = MagicMock()
        mock_get.return_value = mock_response

        from .geocoder import geocode_address

        result = geocode_address("Main St Budapest")
        self.assertEqual(result, (47.5, 19.08))

    @patch("planner.geocoder.time")
    @patch("planner.geocoder.requests.get")
    def test_geocode_not_found(self, mock_get, mock_time):
        mock_time.monotonic.return_value = 100.0
        mock_response = MagicMock()
        mock_response.json.return_value = []
        mock_response.raise_for_status = MagicMock()
        mock_get.return_value = mock_response

        from .geocoder import geocode_address

        result = geocode_address("Nonexistent place xyz")
        self.assertIsNone(result)

    @patch("planner.geocoder.time")
    @patch("planner.geocoder.requests.get")
    def test_geocode_request_failure(self, mock_get, mock_time):
        import requests

        mock_time.monotonic.return_value = 100.0
        mock_get.side_effect = requests.RequestException("Connection error")

        from .geocoder import geocode_address

        result = geocode_address("Some address")
        self.assertIsNone(result)


class OptimizerTest(TestCase):
    def _make_stop(self, stop_id, lat, lng):
        """Create a mock stop object."""
        stop = MagicMock()
        stop.id = stop_id
        stop.lat = lat
        stop.lng = lng
        return stop

    @patch("planner.optimizer.requests.post")
    def test_optimize_route_success(self, mock_post):
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "routes": [
                {
                    "steps": [
                        {"type": "start"},
                        {"type": "job", "job": 2},
                        {"type": "job", "job": 1},
                        {"type": "end"},
                    ]
                }
            ]
        }
        mock_response.raise_for_status = MagicMock()
        mock_post.return_value = mock_response

        from .optimizer import optimize_route

        stops = [self._make_stop(1, 47.5, 19.08), self._make_stop(2, 47.51, 19.09)]
        result = optimize_route(stops)
        self.assertEqual(result, [2, 1])

    @patch("planner.optimizer.requests.post")
    def test_get_route_details_success(self, mock_post):
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "features": [
                {
                    "geometry": {"type": "LineString", "coordinates": [[19.08, 47.5], [19.09, 47.51]]},
                    "properties": {
                        "segments": [
                            {"duration": 120.4, "distance": 1500.7},
                        ]
                    },
                }
            ]
        }
        mock_response.raise_for_status = MagicMock()
        mock_post.return_value = mock_response

        from .optimizer import get_route_details

        stops = [self._make_stop(1, 47.5, 19.08), self._make_stop(2, 47.51, 19.09)]
        result = get_route_details(stops)
        self.assertIsNotNone(result)
        self.assertEqual(result["total_duration"], 120)
        self.assertEqual(result["total_distance"], 1501)
        self.assertEqual(result["geometry"]["type"], "LineString")
        self.assertEqual(len(result["segments"]), 1)

    def test_optimize_fewer_than_2_stops(self):
        from .optimizer import optimize_route

        stop = self._make_stop(42, 47.5, 19.08)
        result = optimize_route([stop])
        self.assertEqual(result, [42])


class OptimizeAPIKeyTest(TestCase):
    """Test that missing or invalid ORS_API_KEY gives clear user-facing errors."""

    def setUp(self):
        self.user, self.client = _make_biker("ors_biker")
        self.session = _make_optimized_session(self.user, num_stops=3)

    @patch("planner.views.sessions.django_settings")
    def test_optimize_missing_api_key(self, mock_settings):
        mock_settings.ORS_API_KEY = ""
        mock_settings.E2E_MOCK = False
        response = self.client.post(f"/api/sessions/{self.session.id}/optimize/")
        self.assertEqual(response.status_code, 503)
        error = response.json()["error"]
        self.assertIn("ORS_API_KEY", error)
        self.assertIn("not configured", error)

    @patch("planner.views.sessions.optimize_route")
    def test_optimize_invalid_api_key(self, mock_optimize):
        mock_response = MagicMock()
        mock_response.status_code = 403
        import requests

        mock_optimize.side_effect = requests.HTTPError("403 Forbidden", response=mock_response)
        response = self.client.post(f"/api/sessions/{self.session.id}/optimize/")
        self.assertEqual(response.status_code, 502)
        error = response.json()["error"]
        self.assertIn("ORS_API_KEY", error)
        self.assertIn("invalid or expired", error)

    @patch("planner.views.sessions.optimize_route")
    def test_optimize_unauthorized_api_key(self, mock_optimize):
        mock_response = MagicMock()
        mock_response.status_code = 401
        import requests

        mock_optimize.side_effect = requests.HTTPError("401 Unauthorized", response=mock_response)
        response = self.client.post(f"/api/sessions/{self.session.id}/optimize/")
        self.assertEqual(response.status_code, 502)
        error = response.json()["error"]
        self.assertIn("ORS_API_KEY", error)


# ============================================
# Priority 7: Serializer Computed Field Tests
# ============================================


class SessionListSerializerTest(TestCase):
    def setUp(self):
        self.user, _ = _make_biker("serial_biker")
        self.session = _make_optimized_session(self.user, num_stops=3)

    def test_delivered_count(self):
        self.session.stops.filter(sequence_order=1).update(delivery_status="delivered")
        self.session.stops.filter(sequence_order=2).update(delivery_status="delivered")
        data = SessionListSerializer(self.session).data
        self.assertEqual(data["delivered_count"], 2)

    def test_not_received_count(self):
        self.session.stops.filter(sequence_order=1).update(delivery_status="not_received")
        data = SessionListSerializer(self.session).data
        self.assertEqual(data["not_received_count"], 1)

    def test_current_stop_name(self):
        self.session.current_stop_index = 2
        self.session.save(update_fields=["current_stop_index"])
        data = SessionListSerializer(self.session).data
        self.assertEqual(data["current_stop_name"], "Stop 2")

    def test_current_stop_name_none(self):
        self.session.current_stop_index = None
        self.session.save(update_fields=["current_stop_index"])
        data = SessionListSerializer(self.session).data
        self.assertIsNone(data["current_stop_name"])


class ActiveSessionSerializerTest(TestCase):
    def setUp(self):
        self.user, _ = _make_biker("active_serial_biker")
        self.session = _make_optimized_session(self.user, num_stops=4)

    def test_delivered_count_includes_all_done(self):
        """ActiveSessionSerializer counts delivered + not_received + skipped."""
        self.session.stops.filter(sequence_order=1).update(delivery_status="delivered")
        self.session.stops.filter(sequence_order=2).update(delivery_status="not_received")
        self.session.stops.filter(sequence_order=3).update(delivery_status="skipped")
        data = ActiveSessionSerializer(self.session).data
        self.assertEqual(data["delivered_count"], 3)


# ============================================
# Priority 8: Live Tracking Endpoint Tests
# ============================================


class LiveTrackingTest(TestCase):
    def setUp(self):
        self.planner, self.planner_client = _make_planner("live_planner")
        self.biker, self.biker_client = _make_biker("live_biker")

    def test_active_sessions_returns_in_progress(self):
        DeliverySession.objects.create(owner=self.biker, name="Active", status="in_progress")
        DeliverySession.objects.create(owner=self.biker, name="Done", status="finished")
        DeliverySession.objects.create(owner=self.biker, name="Waiting", status="not_started")
        response = self.planner_client.get("/api/sessions/active/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 1)
        self.assertEqual(response.json()[0]["name"], "Active")

    def test_active_sessions_biker_rejected(self):
        response = self.biker_client.get("/api/sessions/active/")
        self.assertEqual(response.status_code, 403)

    def test_list_bikers(self):
        _make_biker("extra_biker1")
        _make_biker("extra_biker2")
        response = self.planner_client.get("/api/users/bikers/")
        self.assertEqual(response.status_code, 200)
        usernames = [u["username"] for u in response.json()]
        self.assertIn("live_biker", usernames)
        self.assertIn("extra_biker1", usernames)
        self.assertIn("extra_biker2", usernames)
        self.assertNotIn("live_planner", usernames)


# ============================================
# Priority 9: Fixed Existing Tests
# ============================================


class ParserCSVTest(TestCase):
    def test_parse_csv_with_addresses(self):
        content = b"name,address\nShop A,Main Street 1 Budapest\nShop B,Vaci ut 5 Budapest\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Shop A")
        self.assertEqual(rows[0]["address"], "Main Street 1 Budapest")
        self.assertIsNone(rows[0]["lat"])

    def test_parse_csv_with_coordinates(self):
        content = b"name,address,lat,lng\nDepot,,47.5,19.08\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["lat"], 47.5)
        self.assertEqual(rows[0]["lng"], 19.08)

    def test_parse_csv_mixed_input(self):
        content = b"name,address,lat,lng\nDepot,,47.5,19.08\nShop,Main St,,\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 2)
        self.assertIsNotNone(rows[0]["lat"])
        self.assertIsNone(rows[1]["lat"])
        self.assertEqual(rows[1]["address"], "Main St")

    def test_parse_csv_name_falls_back_to_address(self):
        # A row with an address but no explicit name is a valid deliverable
        # stop (real operational sheets often omit a name column): the name
        # falls back to the address rather than dropping the row.
        content = b"name,address\n,Main St\nShop B,Vaci ut\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Main St")
        self.assertEqual(rows[1]["name"], "Shop B")

    def test_parse_csv_skips_rows_without_name_or_address(self):
        # With neither name, address, nor coordinates there is nothing to
        # deliver to, so the row is dropped.
        content = b"name,address\n,\nShop B,Vaci ut\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Shop B")

    def test_parse_csv_skips_rows_without_address_or_coords(self):
        content = b"name,address,lat,lng\nShop A,,,\nShop B,Main St,,\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)

    def test_parse_csv_bom_handling(self):
        content = b"\xef\xbb\xbfname,address\nShop A,Main St\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Shop A")

    def test_parse_csv_with_delivery_details(self):
        content = b"name,address,lat,lng,product_code,recipient_name,recipient_phone\nShop A,Main St,,,PKG-001,John Doe,+36 30 123 4567\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product_code"], "PKG-001")
        self.assertEqual(rows[0]["recipient_name"], "John Doe")
        self.assertEqual(rows[0]["recipient_phone"], "+36 30 123 4567")

    def test_parse_csv_optional_delivery_details(self):
        content = b"name,address\nShop A,Main St\n"
        rows = parse_csv(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product_code"], "")
        self.assertEqual(rows[0]["recipient_name"], "")
        self.assertEqual(rows[0]["recipient_phone"], "")


class ParserTXTTest(TestCase):
    def test_parse_txt_tab_delimited(self):
        # Fixed: removed dead variable with typo header
        content = b"name\taddress\tlat\tlng\nShop A\tMain St\t\t\n"
        rows = parse_txt(io.BytesIO(content))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Shop A")


class ParserXMLTest(TestCase):
    def test_parse_xml(self):
        content = b"""<deliveries>
            <stop><name>Shop A</name><address>Main St</address><lat></lat><lng></lng></stop>
            <stop><name>Depot</name><address></address><lat>47.5</lat><lng>19.08</lng></stop>
        </deliveries>"""
        rows = parse_xml(io.BytesIO(content))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "Shop A")
        self.assertEqual(rows[1]["lat"], 47.5)


class ParserDispatchTest(TestCase):
    def test_unsupported_format_raises(self):
        with self.assertRaises(ValueError):
            parse_file(io.BytesIO(b"data"), "file.pdf")

    def test_dispatch_csv(self):
        content = b"name,address\nShop,Main St\n"
        rows = parse_file(io.BytesIO(content), "stops.csv")
        self.assertEqual(len(rows), 1)


class SampleDataTest(TestCase):
    """Verify the bundled sample files parse correctly."""

    def _sample_path(self, filename):
        return os.path.join(os.path.dirname(__file__), "sample_data", filename)

    def test_sample_csv(self):
        with open(self._sample_path("budapest_deliveries.csv"), "rb") as f:
            rows = parse_csv(f)
        self.assertEqual(len(rows), 12)
        with_coords = [r for r in rows if r["lat"] is not None]
        self.assertEqual(len(with_coords), 3)

    def test_sample_txt(self):
        with open(self._sample_path("sample.txt"), "rb") as f:
            rows = parse_txt(f)
        self.assertEqual(len(rows), 5)

    def test_sample_xml(self):
        with open(self._sample_path("sample.xml"), "rb") as f:
            rows = parse_xml(f)
        self.assertEqual(len(rows), 5)


class UploadAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="testbiker")
        UserProfile.objects.create(user=self.user, role="biker")
        token = Token.objects.create(user=self.user)
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")

    def test_upload_csv(self):
        content = b"name,address\nShop A,Main St Budapest\nShop B,Vaci ut Budapest\n"
        f = SimpleUploadedFile("stops.csv", content, content_type="text/csv")
        response = self.client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertIn("id", data)
        self.assertEqual(len(data["stops"]), 2)
        self.assertTrue(data["needs_geocoding"])

    def test_upload_no_file(self):
        response = self.client.post("/api/upload/", {}, format="multipart")
        self.assertEqual(response.status_code, 400)

    def test_upload_unsupported_format(self):
        f = SimpleUploadedFile("data.pdf", b"fake", content_type="application/pdf")
        response = self.client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 400)

    def test_upload_csv_with_delivery_details(self):
        content = (
            b"name,address,product_code,recipient_name,recipient_phone\nShop A,Main St,PKG-001,John Doe,+36301234567\n"
        )
        f = SimpleUploadedFile("stops.csv", content, content_type="text/csv")
        response = self.client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 201)
        stop = response.json()["stops"][0]
        self.assertEqual(stop["product_code"], "PKG-001")
        self.assertEqual(stop["recipient_name"], "John Doe")
        self.assertEqual(stop["recipient_phone"], "+36301234567")

    def test_upload_empty_file(self):
        f = SimpleUploadedFile("stops.csv", b"name,address\n", content_type="text/csv")
        response = self.client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 400)

    def test_unauthenticated_upload_rejected(self):
        anon_client = APIClient()
        f = SimpleUploadedFile("stops.csv", b"name,address\nShop,Main St\n", content_type="text/csv")
        response = anon_client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(response.status_code, 401)


class SessionAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="testbiker2")
        UserProfile.objects.create(user=self.user, role="biker")
        token = Token.objects.create(user=self.user)
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")
        self.session = DeliverySession.objects.create(owner=self.user)
        DeliveryStop.objects.create(
            session=self.session, name="Shop A", raw_address="Main St", geocode_status="pending"
        )
        DeliveryStop.objects.create(session=self.session, name="Depot", lat=47.5, lng=19.08, geocode_status="skipped")

    def test_get_session(self):
        response = self.client.get(f"/api/sessions/{self.session.id}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data["stops"]), 2)
        self.assertTrue(data["needs_geocoding"])

    def test_get_session_not_found(self):
        response = self.client.get("/api/sessions/00000000-0000-0000-0000-000000000000/")
        self.assertEqual(response.status_code, 404)

    def test_optimize_needs_minimum_stops(self):
        DeliveryStop.objects.filter(session=self.session, name="Shop A").update(lat=None, lng=None)
        response = self.client.post(f"/api/sessions/{self.session.id}/optimize/")
        self.assertEqual(response.status_code, 400)


# ============================================
# Full Integration Test: Upload → Cluster → Optimize → Deliver
# ============================================


class FullLifecycleIntegrationTest(TestCase):
    """End-to-end test covering the complete multi-route delivery workflow.

    Flow: planner uploads pre-geocoded CSV → clusters into sub-routes →
    moves a stop → optimizes a sub-route → assigns to biker →
    biker starts route → marks all stops delivered → route finishes.
    Also tests uncluster on a separate split session.
    """

    def setUp(self):
        self.planner, self.planner_client = _make_planner("integ_planner")
        self.biker, self.biker_client = _make_biker("integ_biker")

    def _build_csv(self, num_stops=60):
        """Build a pre-geocoded CSV with stops scattered around Budapest."""
        import random

        rng = random.Random(42)
        lines = ["name,lat,lng"]
        centers = [(47.497, 19.040), (47.510, 19.080), (47.480, 19.060)]
        for i in range(num_stops):
            c = centers[i % 3]
            lat = c[0] + rng.uniform(-0.02, 0.02)
            lng = c[1] + rng.uniform(-0.02, 0.02)
            lines.append(f"Stop {i + 1},{lat:.6f},{lng:.6f}")
        return ("\n".join(lines) + "\n").encode()

    def test_full_cluster_optimize_deliver_lifecycle(self):
        # ── Step 1: Planner uploads a 60-stop pre-geocoded CSV ──
        csv_content = self._build_csv(60)
        f = SimpleUploadedFile("cluster_test.csv", csv_content, content_type="text/csv")
        resp = self.planner_client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 201)
        session_id = resp.json()["id"]
        stops = resp.json()["stops"]
        self.assertEqual(len(stops), 60)
        # Pre-geocoded stops should be skipped (not pending)
        self.assertFalse(resp.json()["needs_geocoding"])

        # ── Step 2: Verify session appears in planner's session list ──
        resp = self.planner_client.get("/api/sessions/")
        self.assertEqual(resp.status_code, 200)
        session_ids = [s["id"] for s in resp.json()]
        self.assertIn(session_id, session_ids)

        # ── Step 3: Planner clusters the session into sub-routes ──
        resp = self.planner_client.post(f"/api/sessions/{session_id}/cluster/", {"n_routes": 3})
        self.assertEqual(resp.status_code, 201)
        cluster_data = resp.json()
        self.assertEqual(cluster_data["parent_id"], session_id)
        self.assertEqual(len(cluster_data["sub_routes"]), 3)
        self.assertEqual(cluster_data["cluster_summary"]["n_routes"], 3)
        self.assertEqual(cluster_data["cluster_summary"]["total_stops"], 60)

        sub_route_ids = [sr["id"] for sr in cluster_data["sub_routes"]]
        total_stops_across_routes = sum(sr["stop_count"] for sr in cluster_data["sub_routes"])
        self.assertEqual(total_stops_across_routes, 60)

        # Parent session should now be "split"
        resp = self.planner_client.get(f"/api/sessions/{session_id}/")
        self.assertEqual(resp.json()["status"], "split")

        # Split parents should be hidden from the planner session list too —
        # the sub-routes take their place in the dashboard.
        resp = self.planner_client.get("/api/sessions/")
        planner_ids = [s["id"] for s in resp.json()]
        self.assertNotIn(session_id, planner_ids)
        for sr_id in sub_route_ids:
            self.assertIn(sr_id, planner_ids)

        # Sub-route names follow the "{parent_name}_N" convention.
        sub_routes_in_list = [s for s in resp.json() if s["id"] in sub_route_ids]
        for sr in sub_routes_in_list:
            self.assertRegex(sr["name"], r"_\d+$")

        # ── Step 4: Verify sub-routes are accessible ──
        for sr_id in sub_route_ids:
            resp = self.planner_client.get(f"/api/sessions/{sr_id}/")
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.json()["status"], "not_started")

        # ── Step 5: Move a stop from route 1 to route 2 ──
        route1_id = sub_route_ids[0]
        route2_id = sub_route_ids[1]
        resp = self.planner_client.get(f"/api/sessions/{route1_id}/")
        route1_stops = resp.json()["stops"]
        stop_to_move = route1_stops[0]["id"]
        route1_count_before = len(route1_stops)

        resp = self.planner_client.post(
            f"/api/sessions/{route1_id}/move-stop/",
            {"stop_id": stop_to_move, "to_session_id": route2_id},
        )
        self.assertEqual(resp.status_code, 200)
        move_data = resp.json()
        self.assertEqual(move_data["from_count"], route1_count_before - 1)

        # ── Step 6: Optimize a sub-route (mock ORS) ──
        target_route_id = sub_route_ids[2]
        resp = self.planner_client.get(f"/api/sessions/{target_route_id}/")
        target_stops = resp.json()["stops"]
        target_stop_ids = [s["id"] for s in target_stops]

        with (
            patch("planner.views.sessions.optimize_route") as mock_opt,
            patch("planner.views.sessions.get_route_details") as mock_details,
        ):
            # Mock optimize to return stops in original order
            mock_opt.return_value = target_stop_ids
            mock_details.return_value = {
                "total_duration": 3600,
                "total_distance": 15000,
                "geometry": {"type": "LineString", "coordinates": [[19.04, 47.5], [19.08, 47.51]]},
                "segments": [
                    {"from_index": i, "to_index": i + 1, "duration": 300, "distance": 1200}
                    for i in range(len(target_stops) - 1)
                ],
            }

            resp = self.planner_client.post(f"/api/sessions/{target_route_id}/optimize/")
            self.assertEqual(resp.status_code, 200)
            opt_data = resp.json()
            self.assertIsNotNone(opt_data["route_geometry"])
            self.assertEqual(opt_data["total_duration"], 3600)
            self.assertEqual(opt_data["total_distance"], 15000)

            # Verify stops got sequence_order
            for stop in opt_data["optimized_stops"]:
                self.assertIsNotNone(stop["sequence_order"])

        # ── Step 7: Assign the optimized sub-route to a biker ──
        resp = self.planner_client.patch(
            f"/api/sessions/{target_route_id}/assign/",
            {"owner_id": self.biker.id},
        )
        self.assertEqual(resp.status_code, 200)

        # ── Step 8: Biker sees the assigned route ──
        resp = self.biker_client.get("/api/sessions/")
        biker_session_ids = [s["id"] for s in resp.json()]
        self.assertIn(target_route_id, biker_session_ids)
        # Biker should NOT see the split parent
        self.assertNotIn(session_id, biker_session_ids)

        # ── Step 9: Biker starts the route ──
        resp = self.biker_client.patch(f"/api/sessions/{target_route_id}/start/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["status"], "in_progress")

        # Verify it shows up in active sessions (planner endpoint)
        resp = self.planner_client.get("/api/sessions/active/")
        self.assertEqual(resp.status_code, 200)
        active_ids = [s["id"] for s in resp.json()]
        self.assertIn(target_route_id, active_ids)

        # ── Step 10: Biker delivers all stops ──
        resp = self.biker_client.get(f"/api/sessions/{target_route_id}/")
        route_stops = resp.json()["stops"]
        ordered = sorted([s for s in route_stops if s["sequence_order"]], key=lambda s: s["sequence_order"])

        for i, stop in enumerate(ordered):
            status_choice = "delivered" if i % 3 != 2 else "not_received"
            resp = self.biker_client.patch(
                f"/api/sessions/{target_route_id}/stops/{stop['id']}/status/",
                {"status": status_choice},
            )
            self.assertEqual(resp.status_code, 200)

        # ── Step 11: Verify route auto-finished ──
        resp = self.biker_client.get(f"/api/sessions/{target_route_id}/")
        self.assertEqual(resp.json()["status"], "finished")

        # Active sessions should no longer include this route
        resp = self.planner_client.get("/api/sessions/active/")
        active_ids = [s["id"] for s in resp.json()]
        self.assertNotIn(target_route_id, active_ids)

    def test_uncluster_lifecycle(self):
        """Test the full cluster → uncluster flow."""
        csv_content = self._build_csv(60)
        f = SimpleUploadedFile("uncluster_test.csv", csv_content, content_type="text/csv")
        resp = self.planner_client.post("/api/upload/", {"file": f}, format="multipart")
        session_id = resp.json()["id"]

        # Cluster
        resp = self.planner_client.post(f"/api/sessions/{session_id}/cluster/", {"n_routes": 2})
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.json()["sub_routes"]), 2)

        # Uncluster
        resp = self.planner_client.delete(f"/api/sessions/{session_id}/uncluster/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["deleted_routes"], 2)

        # Parent should be back to not_started
        resp = self.planner_client.get(f"/api/sessions/{session_id}/")
        self.assertEqual(resp.json()["status"], "not_started")
        # Original stops should still be there
        self.assertEqual(len(resp.json()["stops"]), 60)

    def test_uncluster_blocked_by_in_progress(self):
        """Cannot uncluster when a sub-route is in progress."""
        csv_content = self._build_csv(60)
        f = SimpleUploadedFile("block_test.csv", csv_content, content_type="text/csv")
        resp = self.planner_client.post("/api/upload/", {"file": f}, format="multipart")
        session_id = resp.json()["id"]

        resp = self.planner_client.post(f"/api/sessions/{session_id}/cluster/", {"n_routes": 2})
        sub_id = resp.json()["sub_routes"][0]["id"]

        # Start a sub-route: need to assign, optimize, then start
        self.planner_client.patch(f"/api/sessions/{sub_id}/assign/", {"owner_id": self.biker.id})
        # Create optimized stops directly for simplicity
        sub_session = DeliverySession.objects.get(id=sub_id)
        for i, stop in enumerate(sub_session.stops.all(), start=1):
            stop.sequence_order = i
            stop.save(update_fields=["sequence_order"])

        self.biker_client.patch(f"/api/sessions/{sub_id}/start/")

        # Uncluster should be blocked
        resp = self.planner_client.delete(f"/api/sessions/{session_id}/uncluster/")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("in progress", resp.json()["error"].lower())

    def test_planner_upload_creates_unassigned_session(self):
        """When a planner uploads without specifying a biker, the session should be unassigned."""
        csv_content = b"name,address\nShop A,Main St\nShop B,Vaci ut\n"
        f = SimpleUploadedFile("planner_upload.csv", csv_content, content_type="text/csv")
        resp = self.planner_client.post("/api/upload/", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 201)
        self.assertIsNone(resp.json()["owner_name"])

    def test_planner_upload_for_biker(self):
        """When a planner specifies owner_id, the session is assigned to that biker."""
        csv_content = b"name,address\nShop A,Main St\nShop B,Vaci ut\n"
        f = SimpleUploadedFile("biker_upload.csv", csv_content, content_type="text/csv")
        resp = self.planner_client.post("/api/upload/", {"file": f, "owner_id": self.biker.id}, format="multipart")
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()["owner_name"], "integ_biker")


# ============================================
# Split Planner: districts, weighted assignments, remove stop
# ============================================


class DistrictExtractionTest(TestCase):
    def test_budapest_zip_maps_to_district(self):
        from .districts import extract_district

        self.assertEqual(extract_district("1052 Budapest, Vaci utca 10"), 5)
        self.assertEqual(extract_district("Budapest, Bartok Bela ut 1, 1114"), 11)
        self.assertEqual(extract_district("1231 Budapest, Xy utca 2"), 23)

    def test_non_budapest_or_missing_zip(self):
        from .districts import extract_district

        self.assertIsNone(extract_district("6000 Kecskemet, Fo ter 1"))
        self.assertIsNone(extract_district("Vaci utca 10"))
        self.assertIsNone(extract_district(""))
        self.assertIsNone(extract_district(None))
        # 1240+ would be district 24 -> invalid
        self.assertIsNone(extract_district("1245 Budapest"))

    def test_district_label(self):
        from .districts import district_label

        self.assertEqual(district_label(5), "V. kerület")
        self.assertEqual(district_label(13), "XIII. kerület")
        self.assertIsNone(district_label(None))
        self.assertIsNone(district_label(0))


class SplitCountsTest(TestCase):
    def test_proportional_split_sums_to_total(self):
        from .clustering import split_counts

        self.assertEqual(sum(split_counts(50, [30, 15])), 50)
        self.assertEqual(split_counts(10, [1, 1]), [5, 5])
        self.assertEqual(split_counts(0, [1, 2]), [0, 0])

    def test_zero_weights_fall_back_to_even(self):
        from .clustering import split_counts

        self.assertEqual(sum(split_counts(9, [0, 0, 0])), 9)


class SplitPlannerAPITest(TestCase):
    """Cluster endpoint with per-biker assignments, districts endpoint, remove stop."""

    def setUp(self):
        self.planner, self.planner_client = _make_planner("split_planner")
        self.biker_a, _ = _make_biker("split_biker_a")
        self.biker_b, _ = _make_biker("split_biker_b")

        self.session = DeliverySession.objects.create(name="Big Route", status="not_started")
        stops = []
        # 20 stops in district 5 (around one point), 10 stops in district 13 (around another)
        for i in range(20):
            stops.append(
                DeliveryStop(
                    session=self.session,
                    name=f"D5 Stop {i}",
                    raw_address=f"1052 Budapest, Utca {i}",
                    lat=47.49 + i * 0.001,
                    lng=19.05 + i * 0.001,
                    geocode_status="success",
                )
            )
        for i in range(10):
            stops.append(
                DeliveryStop(
                    session=self.session,
                    name=f"D13 Stop {i}",
                    raw_address=f"1136 Budapest, Utca {i}",
                    lat=47.53 + i * 0.001,
                    lng=19.07 + i * 0.001,
                    geocode_status="success",
                )
            )
        DeliveryStop.objects.bulk_create(stops)

    def test_districts_endpoint(self):
        resp = self.planner_client.get(f"/api/sessions/{self.session.id}/districts/")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data["total_stops"], 30)
        self.assertEqual(data["unknown_district_stops"], 0)
        districts = {d["district"]: d["stop_count"] for d in data["districts"]}
        self.assertEqual(districts, {5: 20, 13: 10})

    def test_cluster_with_weighted_assignments(self):
        resp = self.planner_client.post(
            f"/api/sessions/{self.session.id}/cluster/",
            {
                "assignments": [
                    {"biker_id": self.biker_a.id, "target_stops": 20, "district": None},
                    {"biker_id": self.biker_b.id, "target_stops": 10, "district": None},
                ]
            },
            format="json",
        )
        self.assertEqual(resp.status_code, 201)
        sub_routes = resp.json()["sub_routes"]
        self.assertEqual(len(sub_routes), 2)
        by_owner = {sr["owner_name"]: sr["stop_count"] for sr in sub_routes}
        self.assertEqual(by_owner["split_biker_a"], 20)
        self.assertEqual(by_owner["split_biker_b"], 10)

    def test_cluster_with_district_lock(self):
        resp = self.planner_client.post(
            f"/api/sessions/{self.session.id}/cluster/",
            {
                "assignments": [
                    {"biker_id": self.biker_a.id, "target_stops": 20, "district": None},
                    {"biker_id": self.biker_b.id, "target_stops": 10, "district": 13},
                ]
            },
            format="json",
        )
        self.assertEqual(resp.status_code, 201)
        sub_routes = resp.json()["sub_routes"]
        by_owner = {sr["owner_name"]: sr["stop_count"] for sr in sub_routes}
        # Biker B gets exactly the 10 district-13 stops; biker A gets everything else.
        self.assertEqual(by_owner["split_biker_b"], 10)
        self.assertEqual(by_owner["split_biker_a"], 20)

        b_route_id = next(sr["id"] for sr in sub_routes if sr["owner_name"] == "split_biker_b")
        b_route = DeliverySession.objects.get(id=b_route_id)
        self.assertTrue(all(s.raw_address.startswith("1136") for s in b_route.stops.all()))

    def test_cluster_all_district_locked_creates_leftover_route(self):
        resp = self.planner_client.post(
            f"/api/sessions/{self.session.id}/cluster/",
            {
                "assignments": [
                    {"biker_id": self.biker_b.id, "target_stops": 10, "district": 13},
                ]
            },
            format="json",
        )
        self.assertEqual(resp.status_code, 201)
        sub_routes = resp.json()["sub_routes"]
        self.assertEqual(len(sub_routes), 2)
        unassigned = [sr for sr in sub_routes if sr["owner_name"] is None]
        self.assertEqual(len(unassigned), 1)
        self.assertEqual(unassigned[0]["stop_count"], 20)

    def test_cluster_with_unknown_biker_rejected(self):
        resp = self.planner_client.post(
            f"/api/sessions/{self.session.id}/cluster/",
            {"assignments": [{"biker_id": 99999, "target_stops": 30, "district": None}]},
            format="json",
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("not found", resp.json()["error"].lower())

    def test_cluster_with_invalid_district_rejected(self):
        resp = self.planner_client.post(
            f"/api/sessions/{self.session.id}/cluster/",
            {"assignments": [{"biker_id": self.biker_a.id, "target_stops": 30, "district": 42}]},
            format="json",
        )
        self.assertEqual(resp.status_code, 400)

    def test_remove_stop(self):
        stop = self.session.stops.first()
        self.session.total_duration = 1000
        self.session.save(update_fields=["total_duration"])

        resp = self.planner_client.delete(f"/api/sessions/{self.session.id}/stops/{stop.id}/remove/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["remaining_stops"], 29)
        self.assertFalse(self.session.stops.filter(id=stop.id).exists())
        self.session.refresh_from_db()
        self.assertIsNone(self.session.total_duration)

    def test_remove_stop_blocked_in_progress(self):
        self.session.status = "in_progress"
        self.session.save(update_fields=["status"])
        stop = self.session.stops.first()
        resp = self.planner_client.delete(f"/api/sessions/{self.session.id}/stops/{stop.id}/remove/")
        self.assertEqual(resp.status_code, 400)

    def test_remove_stop_requires_planner(self):
        _, biker_client = _make_biker("split_biker_c")
        stop = self.session.stops.first()
        resp = biker_client.delete(f"/api/sessions/{self.session.id}/stops/{stop.id}/remove/")
        self.assertEqual(resp.status_code, 403)
