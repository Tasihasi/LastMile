import csv
import io

from defusedxml import ElementTree as ET
from openpyxl import load_workbook

# Maps real-world (notably Hungarian courier) column headers onto the canonical
# field names this app understands. Keyed by the output of `_normalize_header`
# (lowercased, stripped, spaces -> underscores). This is what lets an untouched
# operational file such as the UPS courier route sheet
# (headers: város, irszám, u, hsz, megj, Cím, kör, <courier names>) upload
# without the user having to re-shape it first.
HEADER_ALIASES = {
    # locality
    "város": "city",
    "telepules": "city",
    "település": "city",
    "irszám": "zip",
    "irsz": "zip",
    "iranyitoszam": "zip",
    # street components
    "u": "street",  # "u " (utca, abbreviated) -> "u"
    "utca": "street",
    "közterület": "street",
    "hsz": "house_number",
    "házszám": "house_number",
    "hazszam": "house_number",
    # full address + note
    "cím": "address",
    "cim": "address",
    "megj": "note",
    "megjegyzés": "note",
    "megjegyzes": "note",
    # assignment / round (the courier a stop belongs to)
    "kör": "route",
    "kor": "route",
    # people
    "név": "name",
    "nev": "name",
    "címzett": "recipient_name",
    "cimzett": "recipient_name",
    "telefon": "recipient_phone",
}


def _stringify(value) -> str:
    """Coerce a cell value to a clean string, dropping the .0 on whole-number floats.

    openpyxl reads numeric cells (e.g. the postal code 1011 or a numeric house
    number) as floats, which would otherwise render as "1011.0" inside a
    geocodable address string. Uncalculated Excel formula strings (those
    starting with "=") are treated as empty: real operational sheets often use
    a formula such as `=C3&" "&D3` to build the address column, which is
    meaningless once exported, so we recompose it from the source columns
    instead.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    if text.startswith("="):
        return ""
    return text


def _compose_address(row: dict) -> str:
    """Build a single geocodable address string from whatever location columns exist.

    Prefers an explicit full-address column; otherwise assembles street + house
    number. When a separate city/postal-code column is present (as in the UPS
    courier sheet, where `Cím` is just "Pala utca 6"), the locality is appended
    so Nominatim has enough context to resolve the right city — but only if the
    address does not already name that city, to avoid duplicating it.

    Args:
        row: Raw row keyed by canonical field names.

    Returns:
        A trimmed address string, or "" if no location data is present.
    """
    address = _stringify(row.get("address")).strip()
    if not address:
        street = _stringify(row.get("street")).strip()
        house = _stringify(row.get("house_number")).strip()
        address = " ".join(p for p in (street, house) if p).strip()

    city = (row.get("city") or "").strip()
    zipcode = _stringify(row.get("zip")).strip()
    locality = " ".join(p for p in (zipcode, city) if p).strip()

    if address and locality and (not city or city.lower() not in address.lower()):
        address = f"{address}, {locality}"

    return address


def _normalize_row(row: dict) -> dict:
    """Normalize a parsed row into {name, address, lat, lng, product_code, recipient_name, recipient_phone}.

    Accepts both the app's native columns and aliased real-world columns (see
    `HEADER_ALIASES`). The stop `name` falls back to the free-text note column
    and finally to the composed address, so operational files that carry no
    explicit "name" column (only an address + a `megj` remark) still yield
    usefully labelled, geocodable stops.
    """
    address = _compose_address(row)
    note = _stringify(row.get("note")).strip()
    name = _stringify(row.get("name")).strip() or note or address
    product_code = (row.get("product_code") or "").strip()
    recipient_name = (row.get("recipient_name") or "").strip()
    recipient_phone = (row.get("recipient_phone") or row.get("phone") or "").strip()
    lat = row.get("lat")
    lng = row.get("lng")

    if lat is not None and lat != "":
        try:
            lat = float(lat)
        except (ValueError, TypeError):
            lat = None
    else:
        lat = None

    if lng is not None and lng != "":
        try:
            lng = float(lng)
        except (ValueError, TypeError):
            lng = None
    else:
        lng = None

    return {
        "name": name,
        "address": address,
        "product_code": product_code,
        "recipient_name": recipient_name,
        "recipient_phone": recipient_phone,
        "lat": lat,
        "lng": lng,
    }


def _validate_rows(rows: list[dict]) -> list[dict]:
    """Validate and filter parsed rows. Each row must have a name and either address or coordinates."""
    valid = []
    for row in rows:
        if not row["name"]:
            continue
        has_coords = row["lat"] is not None and row["lng"] is not None
        has_address = bool(row["address"])
        if has_coords or has_address:
            valid.append(row)
    return valid


def _normalize_header(header: str) -> str:
    """Normalize a header to a canonical field name.

    Lowercases, trims, collapses spaces to underscores, then maps known
    real-world aliases (see `HEADER_ALIASES`) onto canonical field names.
    Unknown headers pass through unchanged so native columns keep working.
    """
    norm = header.strip().lower().replace(" ", "_")
    return HEADER_ALIASES.get(norm, norm)


def parse_csv(file) -> list[dict]:
    """Parse a comma-separated upload into normalized stop dicts.

    Strips a UTF-8 BOM if present so spreadsheets exported by Excel parse
    cleanly. Headers are case- and whitespace-insensitive.
    """
    content = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    reader.fieldnames = [_normalize_header(f) for f in reader.fieldnames]
    rows = [_normalize_row(row) for row in reader]
    return _validate_rows(rows)


def parse_txt(file) -> list[dict]:
    """Parse a tab-delimited upload into normalized stop dicts.

    Same header handling as `parse_csv` but uses tab as the delimiter.
    """
    content = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content), delimiter="\t")
    reader.fieldnames = [_normalize_header(f) for f in reader.fieldnames]
    rows = [_normalize_row(row) for row in reader]
    return _validate_rows(rows)


def parse_xlsx(file) -> list[dict]:
    """Parse the active sheet of an XLSX upload into normalized stop dicts.

    Reads the workbook in read-only mode for memory efficiency on large files.
    Only the *first* (active) sheet is parsed; additional sheets are ignored.
    `data_only=True` returns the last value Excel cached for formula cells (real
    sheets often build the address column with a formula like `=C3&" "&D3`); if
    no cache exists, `_compose_address` rebuilds the address from the street and
    house-number columns instead.
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = [_normalize_header(str(h or "")) for h in next(rows_iter)]
    rows = []
    for row_values in rows_iter:
        row_dict = dict(zip(headers, row_values, strict=False))
        rows.append(_normalize_row(row_dict))

    wb.close()
    return _validate_rows(rows)


def parse_xml(file) -> list[dict]:
    """Parse an XML upload (root with `<stop>` children) into normalized stop dicts.

    Uses defusedxml to defend against XXE / billion-laughs payloads from
    user-supplied files.
    """
    content = file.read()
    root = ET.fromstring(content)

    rows = []
    for stop in root.findall("stop"):
        row = {
            "name": (stop.findtext("name") or "").strip(),
            "address": (stop.findtext("address") or "").strip(),
            "product_code": (stop.findtext("product_code") or "").strip(),
            "recipient_name": (stop.findtext("recipient_name") or "").strip(),
            "recipient_phone": (stop.findtext("recipient_phone") or stop.findtext("phone") or "").strip(),
            "lat": (stop.findtext("lat") or "").strip(),
            "lng": (stop.findtext("lng") or "").strip(),
        }
        rows.append(_normalize_row(row))

    return _validate_rows(rows)


PARSERS = {
    "text/csv": parse_csv,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": parse_xlsx,
    "text/plain": parse_txt,
    "text/xml": parse_xml,
    "application/xml": parse_xml,
}

EXTENSION_MAP = {
    ".csv": parse_csv,
    ".xlsx": parse_xlsx,
    ".txt": parse_txt,
    ".xml": parse_xml,
}


def parse_file(file, filename: str) -> list[dict]:
    """Parse an uploaded file based on its extension. Returns list of normalized stop dicts."""
    import os

    ext = os.path.splitext(filename)[1].lower()
    parser = EXTENSION_MAP.get(ext)
    if parser is None:
        raise ValueError(f"Unsupported file format: {ext}. Supported: .csv, .xlsx, .txt, .xml")
    return parser(file)
